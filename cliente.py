from fastapi import FastAPI, HTTPException
from bs4 import BeautifulSoup
import httpx
import pandas as pd
from pydantic import BaseModel
from typing import List, Optional
import asyncio
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

from historico import adicionar_historico, obter_historico

app = FastAPI(title="API de Scraping Amazon")

class ScrapingRequest(BaseModel):
    url: str
    seletores: dict
    nome_arquivo: Optional[str] = None

class ScrapingResult(BaseModel):
    status: str
    mensagem: str
    caminho_arquivo: Optional[str] = None
    total_itens: Optional[int] = None
    total_paginas: Optional[int] = None

# Configura diretório para salvar arquivos
PASTA_DOWNLOADS = "downloads"
if not os.path.exists(PASTA_DOWNLOADS):
    os.makedirs(PASTA_DOWNLOADS)

async def fazer_scraping(url: str, seletores: dict) -> List[dict]:
    async with httpx.AsyncClient() as cliente:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
            }
            
            await asyncio.sleep(1)
            resposta = await cliente.get(url, headers=headers)
            resposta.raise_for_status()
            
            soup = BeautifulSoup(resposta.text, 'html.parser')
            dados = []
            
            # Encontra todos os produtos
            items = soup.select(seletores['container'])
            
            for item in items:
                dados_item = {}
                for campo, seletor in seletores.items():
                    if campo != 'container':
                        elemento = item.select_one(seletor)
                        dados_item[campo] = elemento.text.strip() if elemento else ''
                dados.append(dados_item)
            
            return dados
            
        except httpx.HTTPError as e:
            raise HTTPException(status_code=400, detail=f"Erro ao acessar URL: {str(e)}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro durante o scraping: {str(e)}")

async def formatar_excel(caminho_arquivo: str, dados: List[dict]):
    """
    Formata o arquivo Excel para melhor visualização
    """
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active
    
    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Formata cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Ajusta largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(caminho_arquivo)

async def salvar_excel(dados: List[dict], nome_arquivo: str):
    try:
        df = pd.DataFrame(dados)
        caminho_arquivo = os.path.join(PASTA_DOWNLOADS, nome_arquivo)
        
        # Salva arquivo Excel
        df.to_excel(caminho_arquivo, index=False, engine='openpyxl')
        
        # Formata o arquivo
        await formatar_excel(caminho_arquivo, dados)
        
        return caminho_arquivo
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar arquivo: {str(e)}")

@app.post("/scraping", response_model=ScrapingResult)
async def iniciar_scraping(request: ScrapingRequest):
    try:
        if not request.nome_arquivo:
            request.nome_arquivo = f"dados_scraping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        results = await fazer_scraping(request.url, request.seletores)
        
        caminho_arquivo = await salvar_excel(results, request.nome_arquivo)
        
        # Adicionar ao histórico
        adicionar_historico(request.nome_arquivo, caminho_arquivo, len(results))

        return ScrapingResult(
            status="sucesso",
            mensagem="Processo de scraping concluído",
            caminho_arquivo=caminho_arquivo,
            total_itens=len(results)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/historico")
async def obter_historico_endpoint():
    return obter_historico()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)