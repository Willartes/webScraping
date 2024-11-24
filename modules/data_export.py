from pathlib import Path
import json
import pandas as pd
from typing import List, Dict
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

class DataExporter:
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    async def export_to_json(self, data: List[Dict], filename: str) -> str:
        """Exporta dados para JSON"""
        output_path = self.output_dir / f"{filename}.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump([d.dict() if hasattr(d, 'dict') else d for d in data], 
                     f, ensure_ascii=False, indent=2)
        return str(output_path)
    
    async def export_to_excel(self, data: List[Dict], filename: str) -> str:
        """Exporta dados para Excel com formatação"""
        if not filename.endswith('.xlsx'):
            filename = f"{filename}.xlsx"
            
        output_path = self.output_dir / filename
        
        # Converte dados para DataFrame
        df = pd.DataFrame([d.dict() if hasattr(d, 'dict') else d for d in data])
        
        # Salva como Excel
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Aplica formatação
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Estilo para cabeçalho
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Aplica estilo no cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva com formatação
        wb.save(output_path)
        
        return str(output_path)
    
    async def export_to_csv(self, data: List[Dict], filename: str) -> str:
        """Exporta dados para CSV"""
        output_path = self.output_dir / f"{filename}.csv"
        df = pd.DataFrame([d.dict() if hasattr(d, 'dict') else d for d in data])
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return str(output_path)